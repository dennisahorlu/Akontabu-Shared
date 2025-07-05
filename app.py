from flask import Flask, render_template, request, jsonify, redirect, url_for, session, g, flash
import os
from datetime import datetime, date, timezone
from sqlalchemy import event, Engine
from sqlalchemy.orm import joinedload
from flask_login import current_user, login_required, login_user
from extensions import db, login_manager
from models import User, Product, Sale, Payment
import math
from sqlalchemy import func
from flask import send_file, make_response
from io import BytesIO
import pandas as pd
from weasyprint import HTML
from jinja2 import Environment, FileSystemLoader
from babel.numbers import format_currency
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from xhtml2pdf import pisa
import openpyxl
from openpyxl.utils import get_column_letter
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as ExcelImage

app = Flask(__name__)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///akontabu.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
# app.config['UPLOAD_FOLDER'] = os.path.join('static', 'uploads')
app.config['UPLOAD_FOLDER'] = os.path.join(app.static_folder, 'uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)


# Initialize extensions
db.init_app(app)
login_manager.init_app(app)
login_manager.login_view = 'login'

@app.template_filter('currency')
def format_currency_filter(value, currency=None):
    if not value:
        return ''
    if not currency:
        currency = getattr(current_user, 'currency_code', 'GHS')  # ✅ Correct field
    try:
        return format_currency(value, currency, locale='en')
    except:
        return f"{currency} {value:,.2f}"


# Enable SQLite foreign key constraints
@event.listens_for(Engine, "connect")
def set_sqlite_pragma(dbapi_connection, connection_record):
    cursor = dbapi_connection.cursor()
    cursor.execute("PRAGMA foreign_keys=ON")
    cursor.close()

def round_up_to_half(v): 
    return math.ceil(v * 2) / 2

# Auth middleware
@app.before_request
def load_user():
    if 'user_id' in session:
        g.user = db.session.get(User, session['user_id'])

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))  # assuming you're using SQLAlchemy and User has an id

@app.context_processor
def inject_user():
    return dict(current_user=current_user)


# Routes
@app.route('/')
def index(): 
     return render_template ('index.html')

# Routes
@app.route('/signup', methods=['GET', 'POST'])
def signup():
    if request.method == 'GET':
        return render_template('signup.html')
    
    data = request.get_json()
    if User.query.filter_by(email=data['email']).first():
        return jsonify({'message': 'Email already registered'}), 400
    
    user = User(
        name=data['name'],
        email=data['email'],
        business_name=data.get('business_name', ''),  # Make sure this is included
            
    )
    user.set_password(data['password'])
    db.session.add(user)
    db.session.commit()
    
    return jsonify({'message': 'User created successfully'}), 201

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return render_template('login.html')
    
    try:
        data = request.get_json()
        user = db.session.execute(
    db.select(User).where(User.email == data['email'])
).scalar_one_or_none()
        
        if user and user.check_password(data['password']):
            login_user(user)  # Make sure this is called
            session['user_id'] = user.id
            return jsonify({'message': 'Login successful'}), 200
        
        return jsonify({'message': 'Invalid email or password'}), 401
    except Exception as e:
        return jsonify({'message': 'An error occurred'}), 500

@app.route('/dashboard')
@login_required
def dashboard():
    product_costs = db.session.query(
    func.sum(Payment.total_cost)
    ).filter(
        Payment.payment_type == 'product',
        Payment.user_id == current_user.id
    ).scalar() or 0

    other_costs = db.session.query(
        func.sum(Payment.amount)
    ).filter(
        Payment.payment_type != 'product',
        Payment.user_id == current_user.id
    ).scalar() or 0


    total_sales = db.session.query(
        func.sum(Sale.quantity * Sale.price)
    ).filter(Sale.user_id == current_user.id).scalar() or 0

    total_expenditure = product_costs + other_costs

    page = request.args.get('page', 1, type=int)
    per_page = 5

    # ✅ Only fetch products with quantity > 0
    products_query = Product.query.filter_by(user_id=current_user.id).filter(Product.quantity > 0)
    pagination = products_query.paginate(page=page, per_page=per_page, error_out=False)

    # ✅ Build inventory details from paginated products
    inventory = [{
        'name': p.name,
        'quantity': p.quantity,
        'value': round((p.quantity or 0) * (p.selling_price or 0), 2)
    } for p in pagination.items]

    # ✅ Total inventory value of all available products (not just paginated)
    inventory_value = db.session.query(
        func.sum(Product.quantity * Product.selling_price)
    ).filter(Product.user_id == current_user.id, Product.quantity > 0).scalar() or 0

    return render_template('dashboard.html',
        total_sales=round(total_sales, 2),
        total_expenditure=round(total_expenditure, 2),
        product_costs=round(product_costs, 2),
        other_costs=round(other_costs, 2),
        inventory_value=round(inventory_value, 2),
        inventory=inventory,
        products=pagination.items,
        pagination=pagination
    )

@app.route('/inventory')
@login_required
def inventory():
    page = request.args.get('page', 1, type=int)
    per_page = 5

    # ✅ Only paginate products with quantity > 0
    products_query = Product.query.filter_by(user_id=current_user.id).filter(Product.quantity > 0)
    pagination = products_query.order_by(Product.name).paginate(page=page, per_page=per_page, error_out=False)

    return render_template(
        'inventory.html',
        products=pagination.items,
        pagination=pagination
    )

def update_inventory_from_payment(payment):
    normalized_name = payment.product_name.strip().title()  # 🧠 Normalize the product name

    # Update the payment object too (if not already set)
    payment.product_name = normalized_name

    product = Product.query.filter_by(name=normalized_name, user_id=payment.user_id).first()

    if product:
        # 🧮 Weighted average cost
        old_qty = product.quantity
        old_total_cost = old_qty * product.cost_price
        new_total_cost = payment.quantity * payment.cost_per_unit
        new_qty = old_qty + payment.quantity

        if new_qty > 0:
            average_cost = (old_total_cost + new_total_cost) / new_qty
            product.cost_price = round(average_cost, 2)

        product.quantity = new_qty
        product.selling_price = payment.selling_price
        product.profit_margin = payment.profit_margin

    else:
        product = Product(
            name=normalized_name,
            quantity=payment.quantity,
            cost_price=payment.cost_per_unit,
            selling_price=payment.selling_price,
            profit_margin=payment.profit_margin,
            user_id=payment.user_id
        )
        db.session.add(product)

    db.session.commit()

@app.route('/sales', methods=['GET', 'POST'])
def sales():
    if 'user_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        data = request.get_json()
        product = Product.query.filter_by(id=data['product_id'], user_id=session['user_id']).first()

        if not product:
            return jsonify({'message': 'Product not found'}), 404

        try:
            quantity = int(data['quantity'])
            price = float(data['price'])
        except (ValueError, TypeError):
            return jsonify({'message': 'Invalid quantity or price'}), 400

        if quantity > product.quantity:
            return jsonify({'message': 'Not enough stock available'}), 400

        # Reduce inventory
        product.quantity -= quantity

        # Create new sale
        sale = Sale(
            product_id=product.id,
            quantity=quantity,
            price=price,
            user_id=session['user_id'],
            customer=data.get('customer', ''),
            description=data.get('description', '')
        )

        db.session.add(sale)
        db.session.commit()
        return jsonify({'message': 'Sale recorded successfully'}), 201

    # For GET request - load sales and products
    start = request.args.get('start')
    end = request.args.get('end')

    sales_query = Sale.query.filter_by(user_id=session['user_id'])

    if start and end:
        try:
            start_date = datetime.strptime(start, '%Y-%m-%d')
            end_date = datetime.strptime(end, '%Y-%m-%d')
            sales_query = sales_query.filter(Sale.date.between(start_date, end_date))
        except ValueError:
            flash("Invalid date range", "error")

    sales_query = sales_query.order_by(Sale.date.desc())

    # Apply pagination to sales
    page = request.args.get('page', 1, type=int)
    per_page = 7
    pagination = sales_query.paginate(page=page, per_page=per_page, error_out=False)

    # Load all products (for sale form)
    products = Product.query.filter_by(user_id=session['user_id']).all()

    return render_template('sales.html',
        sales=pagination.items,
        pagination=pagination,
        products=products
    )



@app.route('/payments')
@login_required
def payments():
    query = Payment.query.filter_by(user_id=current_user.id)

    # Optional filters
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')

    if start_date:
        query = query.filter(Payment.date >= datetime.strptime(start_date, "%Y-%m-%d"))
    if end_date:
        query = query.filter(Payment.date <= datetime.strptime(end_date, "%Y-%m-%d"))

    query = query.order_by(Payment.date.desc())

    page = request.args.get('page', 1, type=int)
    per_page = 3

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    payments = pagination.items  # ✅ Use only paginated items

    return render_template(
        'payments.html',
        payments=payments,
        pagination=pagination,
        current_date=date.today().strftime('%Y-%m-%d')
    )

@app.route('/reports')
@login_required
def reports():
    start = request.args.get('start')
    end = request.args.get('end')
    tab = request.args.get('tab', 'profit')  # <-- Track which tab is active

    def date_filter(query, model):
        if start and end:
            try:
                s_date = datetime.strptime(start, '%Y-%m-%d')
                e_date = datetime.strptime(end, '%Y-%m-%d')
                return query.filter(model.date.between(s_date, e_date))
            except ValueError:
                flash("Invalid date range")
        return query

    sales = date_filter(Sale.query.filter_by(user_id=current_user.id), Sale).all()
    payments = date_filter(Payment.query.filter_by(user_id=current_user.id), Payment).all()
    inventory = Product.query.filter_by(user_id=current_user.id).all()

    total_revenue = sum(s.price * s.quantity for s in sales)
    total_payments = sum(p.amount for p in payments)
    sales_total = sum(s.quantity * s.price for s in sales)
    payments_total = sum(p.amount for p in payments)
    profit = total_revenue - total_payments

    inventory_data = [{
        'name': p.name,
        'quantity': p.quantity,
        'cost_price': p.cost_price,
        'selling_price': p.selling_price,
        'value': p.cost_price * p.quantity
    } for p in inventory]

    return render_template(
        'reports.html',
        sales=sales,
        payments=payments,
        inventory_data=inventory_data,
        total_revenue=total_revenue,
        total_payments=total_payments,
        payments_total=payments_total,
        sales_total=sales_total,
        profit=profit,
        start=start,
        end=end,
        now=datetime.now(),
        tab=tab  # <-- Pass to template
    )


@app.route('/reports/sales')
@login_required
def sales_report():
    start = request.args.get('start')
    end = request.args.get('end')

    def date_filter(query, model):
        if start and end:
            try:
                s_date = datetime.strptime(start, '%Y-%m-%d')
                e_date = datetime.strptime(end, '%Y-%m-%d')
                return query.filter(model.date.between(s_date, e_date))
            except ValueError:
                flash("Invalid date range", "error")
        return query

    # Filtered sales
    sales = date_filter(Sale.query.filter_by(user_id=current_user.id), Sale).order_by(Sale.date.asc()).all()

    # Total sales = sum of quantity × price
    sales_total = sum(s.quantity * s.price for s in sales)

    return render_template(
        'reports/partials/sales.html',
        report_title="Sales Report",
        sales=sales,
        sales_total=sales_total,
        start=start,
        end=end,
        now=datetime.now()
    )

@app.route('/reports/payments')
@login_required
def payment_report():
    start = request.args.get('start')
    end = request.args.get('end')
    product_id = request.args.get('product')

    query = Payment.query.filter_by(user_id=current_user.id)

    if start and end:
        try:
            s_date = datetime.strptime(start, '%Y-%m-%d')
            e_date = datetime.strptime(end, '%Y-%m-%d')
            query = query.filter(Payment.date.between(s_date, e_date))
        except ValueError:
            flash("Invalid date format. Use YYYY-MM-DD.")

    if product_id:
        query = query.filter_by(product_id=product_id)

    payments = query.order_by(Payment.date.desc()).all()
    products = Product.query.filter_by(user_id=current_user.id).all()

    total_amount = sum(p.amount for p in payments)

    return render_template(
        'partials/payments.html',
        payments=payments,
        all_products=products,
        start=start,
        end=end,
        payments_total=total_amount
    )


@app.route('/export/sales/excel')
@login_required
def export_sales_excel():
    start = request.args.get('start')
    end = request.args.get('end')

    def date_filter(query, model):
        if start and end:
            try:
                s_date = datetime.strptime(start, '%Y-%m-%d')
                e_date = datetime.strptime(end, '%Y-%m-%d')
                return query.filter(model.date.between(s_date, e_date))
            except ValueError:
                flash("Invalid date range", "error")
        return query

    sales = date_filter(Sale.query.filter_by(user_id=current_user.id), Sale).order_by(Sale.date.asc()).all()
    grand_total = sum(s.quantity * s.price for s in sales)

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # Add logo from upload folder
    from openpyxl.drawing.image import Image as ExcelImage
    import os

    logo_filename = current_user.logo_path or 'default_logo.png'
    logo_full_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)

    if os.path.exists(logo_full_path):
        img = ExcelImage(logo_full_path)
        img.height = 80
        img.width = 160
        ws.add_image(img, 'A1')

    # Header Info
    ws.merge_cells('A1:F1')
    ws['A1'] = current_user.business_name or "AKONTABU"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:F2')
    ws['A2'] = "Sales Report"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A3:F3')
    period_text = f"Period: {start} to {end}" if start and end else "Period: All Time"
    ws['A3'] = period_text
    ws['A3'].font = Font(size=10)
    ws['A3'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A4:F4')
    ws['A4'] = f"Date Generated: {datetime.now().strftime('%Y-%m-%d %I:%M:%S %p')}"
    ws['A4'].alignment = Alignment(horizontal="center")

    # Table Header
    headers = ['Date & Time', 'Product', 'Customer', 'Quantity', 'Price', 'Total']
    header_row = 6
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    # Table Rows
    for idx, s in enumerate(sales, start=header_row + 1):
        ws.cell(row=idx, column=1, value=s.date.strftime('%Y-%m-%d %I:%M:%S %p'))
        ws.cell(row=idx, column=2, value=s.product.name)
        ws.cell(row=idx, column=3, value=getattr(s, 'customer', 'N/A') if isinstance(s.customer, str) else (s.customer.name if s.customer else 'N/A'))
        ws.cell(row=idx, column=4, value=s.quantity)
        ws.cell(row=idx, column=5, value=s.price)
        ws.cell(row=idx, column=6, value=s.quantity * s.price)

    # Format currency and borders
    currency_format = '"GHS" #,##0.00'
    last_data_row = ws.max_row

    for row in ws.iter_rows(min_row=header_row + 1, max_row=last_data_row):
        for i, cell in enumerate(row):
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if i in [4, 5]:  # Price or Total
                cell.number_format = currency_format
                cell.alignment = Alignment(horizontal="right")

    # Grand Total
    total_row = last_data_row + 2
    ws.cell(row=total_row, column=4, value="Grand Total:").font = Font(bold=True)
    ws.cell(row=total_row, column=6, value=grand_total).font = Font(bold=True)
    ws.cell(row=total_row, column=6).number_format = currency_format
    ws.cell(row=total_row, column=6).alignment = Alignment(horizontal="right")

    # Export as file
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, download_name='sales_report.xlsx', as_attachment=True)

@app.route('/export/sales/pdf')
@login_required
def export_sales_pdf():
    start = request.args.get("start")
    end = request.args.get("end")

    def apply_date_filter(query):
        if start and end:
            try:
                start_date = datetime.strptime(start, "%Y-%m-%d")
                end_date = datetime.strptime(end, "%Y-%m-%d")
                return query.filter(Sale.date.between(start_date, end_date))
            except ValueError:
                pass
        return query

    sales = apply_date_filter(Sale.query.filter_by(user_id=current_user.id)).order_by(Sale.date.asc()).all()
    sales_total = sum(s.quantity * s.price for s in sales)
    total_revenue = sales_total
    total_payments = sum(p.amount for p in Payment.query.filter_by(user_id=current_user.id).all())
    profit = total_revenue - total_payments

    # Render your partial template or full report
    html_content = render_template(
        'reports/partials/sales.html',  # Or 'reports/sales_pdf.html' if you want to create a special version
        sales=sales,
        sales_total=sales_total,
        total_revenue=total_revenue,
        total_payments=total_payments,
        profit=profit,
        start=start,
        end=end,
        now=datetime.now(),
        current_user=current_user
    )

    pdf_file = BytesIO()
    HTML(string=html_content, base_url=request.host_url).write_pdf(pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        mimetype='application/pdf',
        download_name='sales_report.pdf',
        as_attachment=True
    )

@app.route('/reports/payments/download/excel')
@login_required
def download_payments_excel():
    start = request.args.get('start')
    end = request.args.get('end')
    payment_type = request.args.get('payment_type')
    product_name = request.args.get('product_name')

    def date_filter(query, model):
        if start and end:
            try:
                s_date = datetime.strptime(start, '%Y-%m-%d')
                e_date = datetime.strptime(end, '%Y-%m-%d')
                return query.filter(model.date.between(s_date, e_date))
            except ValueError:
                flash("Invalid date range")
        return query

    query = Payment.query.filter_by(user_id=current_user.id)
    query = date_filter(query, Payment)

    if payment_type:
        query = query.filter(Payment.payment_type == payment_type)
    if product_name:
        query = query.join(Product).filter(Product.name.ilike(f"%{product_name}%"))

    payments = query.order_by(Payment.date.asc()).all()
    payments_total = sum(p.amount for p in payments)

    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"

    # Add logo
    logo_filename = current_user.logo_path or 'default_logo.png'
    logo_path = os.path.join(app.config['UPLOAD_FOLDER'], logo_filename)
    if os.path.exists(logo_path):
        try:
            img = ExcelImage(logo_path)
            img.height = 80
            img.width = 160
            ws.add_image(img, 'A1')
        except Exception as e:
            print("Image error:", e)

    # Header info
    ws.merge_cells('A1:E1')
    ws['A1'] = current_user.business_name or "AKONTABU"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:E2')
    ws['A2'] = "Payment Report"
    ws['A2'].font = Font(size=12, bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A3:E3')
    period_text = f"Period: {start} to {end}" if start and end else "Period: All Time"
    ws['A3'] = period_text
    ws['A3'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A4:E4')
    ws['A4'] = f"Generated on: {datetime.now().strftime('%d-%m-%Y')}"
    ws['A4'].alignment = Alignment(horizontal="center")

    # Table headers
    headers = ['Date', 'Type', 'Description', 'Amount (GHS)']
    header_row = 6
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # Table data
    for idx, p in enumerate(payments, start=header_row + 1):
        ws.cell(row=idx, column=1, value=p.date.strftime('%d-%m-%Y'))
        ws.cell(row=idx, column=2, value=p.product.name if p.product else p.payment_type)
        ws.cell(row=idx, column=3, value=p.description or '')
        amount_cell = ws.cell(row=idx, column=4, value=p.amount)
        amount_cell.number_format = '"GHS" #,##0.00'
        amount_cell.alignment = Alignment(horizontal='right')

    # Grand Total row
    total_row = ws.max_row + 2
    ws.cell(row=total_row, column=3, value="Grand Total:").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=4, value=payments_total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '"GHS" #,##0.00'
    total_cell.alignment = Alignment(horizontal='right')

    # Report Info sheet (optional)
    info_data = {
        'Business Name': [current_user.business_name or 'AKONTABU'],
        'Report Type': ['Payment Report'],
        'Start Date': [start or 'All Time'],
        'End Date': [end or 'All Time'],
        'Payment Type': [payment_type or 'All'],
        'Product Name': [product_name or 'All'],
        'Generated On': [datetime.now().strftime('%d-%m-%Y')],
        'Total Payments': [f"GHS {round(payments_total, 2)}"]
    }
    info_df = pd.DataFrame(info_data)
    info_ws = wb.create_sheet("Report Info")
    for r, (label, value) in enumerate(info_data.items(), start=1):
        info_ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        info_ws.cell(row=r, column=2, value=value[0])

    # Export to file
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name='payment_report.xlsx', as_attachment=True)

@app.route('/reports/payments/download/pdf')
@login_required
def export_payment_pdf():
    start = request.args.get('start')
    end = request.args.get('end')
    payment_type = request.args.get('payment_type')
    product_name = request.args.get('product_name')

    def date_filter(query):
        if start and end:
            try:
                s_date = datetime.strptime(start, '%Y-%m-%d')
                e_date = datetime.strptime(end, '%Y-%m-%d')
                return query.filter(Payment.date.between(s_date, e_date))
            except ValueError:
                pass
        return query

    # Apply filters
    query = Payment.query.filter_by(user_id=current_user.id)
    query = date_filter(query)

    if payment_type:
        query = query.filter(Payment.payment_type == payment_type)
    if product_name:
        query = query.join(Product).filter(Product.name.ilike(f"%{product_name}%"))

    payments = query.order_by(Payment.date.asc()).all()
    payments_total = sum(p.amount for p in payments)
    

    # Render the HTML
    html_content = render_template(
        'reports/partials/payments.html',
        payments=payments,
        payments_total=payments_total,
        start=start,
        end=end,
        now=datetime.now(),
        report_title="Payment Report",
        current_user=current_user  # required for business name and logo
    )

    # Generate PDF
    pdf_file = BytesIO()
    HTML(string=html_content, base_url=request.host_url).write_pdf(pdf_file)
    pdf_file.seek(0)

    return send_file(
        pdf_file,
        mimetype='application/pdf',
        download_name='payment_report.pdf',
        as_attachment=True
    )

@app.route('/settings')
@login_required
def settings():
    return render_template('settings.html', current_currency=current_user.currency_code)


@app.route('/logout')
def logout():
    session.pop('user_id', None)
    return redirect(url_for('index'))

@app.route('/record-sale', methods=['POST'])
def record_sale():
    product_id = request.form.get('product_id')
    quantity = request.form.get('quantity')
    
    # Validate inputs
    if not product_id or not quantity:
        return jsonify({"error": "Missing data"}), 400
    
    # Save to database (example using SQLAlchemy)
    sale = Sale(product_id=product_id, quantity=quantity)
    db.session.add(sale)
    db.session.commit()
    
    return jsonify({"success": True})

@app.route('/api/products/<int:product_id>')
def get_product(product_id):
    product = Product.query.get(product_id)
    if product:
        return jsonify({'id': product.id, 'name': product.name, 'selling_price': product.selling_price})
    return jsonify({'error': 'Product not found'}), 404

@app.route('/api/payments', methods=['POST'])
@login_required
def record_payment():
    data = request.get_json()

    if not data or 'payment_type' not in data:
        return jsonify({'message': 'Invalid data'}), 400

    if data['payment_type'] == 'product':
        try:
            payment = Payment(
                payment_type='product',
                product_name=data['product_name'].strip().title(),
                quantity=data['quantity'],
                base_cost=data['base_cost'],
                transportation=data['transportation'],
                carriage=data['carriage'],
                total_cost=data['base_cost'] + data['transportation'] + data['carriage'],
                cost_per_unit=(data['base_cost'] + data['transportation'] + data['carriage']) / data['quantity'],
                profit_margin=data['profit_margin'],
                selling_price=data['calculated_price'],
                amount=data['quantity'] * data['calculated_price'],
                user_id=current_user.id,
                date=datetime.now(timezone.utc)
            )
            db.session.add(payment)

            # Update or create product in inventory
            normalized_name = payment.product_name
            existing_product = Product.query.filter_by(name=normalized_name, user_id=current_user.id).first()

            if existing_product:
                existing_product.quantity += payment.quantity
                existing_product.cost_price = payment.cost_per_unit  # Optional: depends on logic
                existing_product.selling_price = payment.selling_price
            else:
                new_product = Product(
                    name=normalized_name,
                    quantity=payment.quantity,
                    cost_price=payment.cost_per_unit,
                    selling_price=payment.selling_price,
                    user_id=current_user.id
                )
                db.session.add(new_product)

            db.session.commit()
            return jsonify({'message': 'Payment recorded successfully'}), 201

        except Exception as e:
            db.session.rollback()
            return jsonify({'message': f'Error: {str(e)}'}), 500

    else:
        try:
            payment = Payment(
                payment_type=data['payment_type'],
                amount=data['amount'],
                description=data['description'].strip().title(),
                user_id=current_user.id,
                date=datetime.now(timezone.utc)
            )
            db.session.add(payment)
            db.session.commit()
            return jsonify({'message': 'Payment recorded successfully'}), 201

        except Exception as e:
            db.session.rollback()
            return jsonify({'message': f'Error: {str(e)}'}), 500


@app.route('/api/payments/<int:payment_id>', methods=['DELETE'])
@login_required
def delete_payment(payment_id):
    payment = Payment.query.get_or_404(payment_id)
    if payment.user_id != current_user.id:
        return jsonify({'message': 'Unauthorized'}), 403

    try:
        db.session.delete(payment)
        db.session.commit()
        return jsonify({'message': 'Deleted successfully'}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Server error: {str(e)}'}), 500
    
    if payment.payment_type == 'product':
        product = Product.query.filter_by(name=payment.product_name, user_id=payment.user_id).first()
        if product:
            product.quantity -= payment.quantity
            if product.quantity <= 0:
                db.session.delete(product)
            else:
                db.session.add(product)

    
@app.route('/api/payments/restore', methods=['POST'])
@login_required
def restore_payment():
    data = request.get_json()
    
    try:
        payment = Payment(
            payment_type=data['payment_type'],
            amount=data['amount'],
            description=data.get('description', ''),
            user_id=current_user.id,
            date=datetime.strptime(data['date'], '%Y-%m-%d'),
            product_name=data.get('product_name'),
            quantity=data.get('quantity'),
            base_cost=data.get('base_cost'),
            transportation=data.get('transportation'),
            carriage=data.get('carriage'),
            total_cost=data.get('total_cost'),
            cost_per_unit=data.get('cost_per_unit'),
            profit_margin=data.get('profit_margin'),
            selling_price=data.get('selling_price'),
            product_id=data.get('product_id')
        )
        db.session.add(payment)
        db.session.commit()

        return jsonify({
            'message': 'Payment restored',
            'payment_id': payment.id,
            'date': payment.date.strftime('%Y-%m-%d')
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error restoring payment: {str(e)}'}), 500
    
@app.route('/api/payments/<int:id>', methods=['PUT'])
@login_required
def update_payment(id):
    payment = Payment.query.get_or_404(id)

    if payment.user_id != current_user.id:
        return jsonify({'message': 'Unauthorized'}), 403

    data = request.get_json()

    # Update common fields
    payment.payment_type = data['payment_type']
    payment.amount = data.get('amount')
    payment.description = data.get('description')

    # If it's a product payment, update product details
    if data['payment_type'] == 'product':
        quantity = data.get('quantity', 1)
        base_cost = data.get('base_cost', 0)
        transportation = data.get('transportation', 0)
        carriage = data.get('carriage', 0)
        total_cost = base_cost + transportation + carriage
        cost_per_unit = total_cost / quantity if quantity else 0
        selling_price = data.get('calculated_price', 0)
        amount = round(quantity * selling_price, 2)

        payment.product_name = data['product_name']
        payment.quantity = quantity
        payment.base_cost = base_cost
        payment.transportation = transportation
        payment.carriage = carriage
        payment.total_cost = total_cost
        payment.cost_per_unit = cost_per_unit
        payment.profit_margin = data.get('profit_margin', 0)
        payment.selling_price = selling_price
        payment.amount = amount


    db.session.commit()

    return jsonify({'message': 'Payment updated successfully'})

@app.route('/api/payments/<int:id>', methods=['GET'])
@login_required
def get_payment(id):
    payment = Payment.query.get_or_404(id)
    
    if payment.user_id != current_user.id:
        return jsonify({'message': 'Unauthorized'}), 403

    return jsonify({
        'id': payment.id,
        'payment_type': payment.payment_type,
        'description': payment.description,
        'amount': payment.amount,
        'product_name': payment.product_name,
        'quantity': payment.quantity,
        'base_cost': payment.base_cost,
        'transportation': payment.transportation,
        'carriage': payment.carriage,
        'total_cost': payment.total_cost,
        'cost_per_unit': payment.cost_per_unit,
        'profit_margin': payment.profit_margin,
        'selling_price': payment.selling_price,
        'date': payment.date.strftime('%Y-%m-%d'),
        'product_id': payment.product_id
    })


@app.route('/settings/currency', methods=['POST'])
@login_required
def update_currency():
    data = request.get_json()
    currency = data.get('currency')
    if currency:
        current_user.currency_code = currency
        db.session.commit()  
        return jsonify({'message': 'Currency updated'}), 200
    return jsonify({'error': 'No currency provided'}), 400

@app.route('/settings/profile', methods=['POST'])
@login_required
def update_profile():
    business_name = request.form.get('business_name')
    email = request.form.get('email')
    password = request.form.get('password')
    logo_file = request.files.get('logo')

    user = current_user
    user.business_name = business_name
    user.email = email

    if password:
        from werkzeug.security import generate_password_hash
        user.password = generate_password_hash(password)

    if logo_file:
        filename = secure_filename(logo_file.filename)
        logo_path = os.path.join('uploads', filename)  # or your folder e.g. static/uploads
        full_path = os.path.join(app.static_folder, 'uploads', filename)
        logo_file.save(full_path)
        user.logo_path = f'uploads/{filename}'

    db.session.commit()

    return jsonify({
        'message': 'Profile updated successfully',
        'logo_url': url_for('static', filename=user.logo_path)
    })





 # Initialize database
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    app.run(debug=True)

    